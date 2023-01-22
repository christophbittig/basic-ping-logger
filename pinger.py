import argparse
from datetime import datetime, timedelta
import re
import subprocess

from openpyxl.chart import LineChart, Reference
from openpyxl import Workbook, load_workbook
import pandas as pd


def run_command(command: str) -> str:
    raw_output = subprocess.run(command, stdout=subprocess.PIPE)
    return raw_output.stdout.decode('utf-8')

def parse_packet_loss(ping_response: str) -> int:
    packet_loss = 0
    if match := re.search(r'\((\d+)% loss\)', ping_response):
        packet_loss = int(match.group(1))

    return packet_loss
    
def parse_latency(ping_response: str) -> list[int]:
    packet_lines:list[str] = ping_response.splitlines()[2:6]

    response_times = []
    for packet in packet_lines:
        time = 0
        if matched_time := re.search(r"time=(\d+)ms", packet):
            time = int(matched_time.group(1))

        response_times.append(time)

    return response_times

def _update_excel_chart(filename: str, num_rows:int) -> None:
    excel_workbook: Workbook = load_workbook(filename)
    worksheet_name = excel_workbook.sheetnames[0]
    worksheet_object = excel_workbook[worksheet_name]

    chart = LineChart()
    chart.height = 20
    chart.width = 40
    chart.title = "Latency over Time"
    chart.style = 15

    data = Reference(worksheet_object, min_col = 3, min_row = 1, max_col = 4, max_row=num_rows)
    chart.add_data(data, titles_from_data=True)

    dropped_packet_series = chart.series[1]
    dropped_packet_series.marker.symbol = "triangle"
    dropped_packet_series.marker.graphicalProperties.solidFill = "FF0000"
    dropped_packet_series.marker.graphicalProperties.line.solidFill = "FF0000"

    dropped_packet_series.graphicalProperties.line.noFill = True

    worksheet_object.add_chart(chart, "E1")

    excel_workbook.save(filename)

def save_data(data: pd.DataFrame, filename: str) -> bool:
    try:
        data.to_excel(filename)
        _update_excel_chart(filename, len(data))
        print(f"{datetime.now()} | INFO: Saved {len(data)} rows of data to {filename} successfully.")
        return True
    except Exception as e:
        print(f"{datetime.now()} | WARNING: Failed to save with exception:\n{e}")
        return False

def main_loop(save_filename:str, target: str, save_frequency: int):
    print(f"Starting ping logger at {datetime.now()}...")
    my_data = pd.DataFrame(columns=['Time', 'Latency', 'Packet Dropped'])
    while True:
        time_when_command_started = datetime.now()
        ping_response: str = run_command(['ping', target])
        response_times: list[int] = parse_latency(ping_response)
        # Not used now but could be used in a rolling average at some point
        packet_loss: int = parse_packet_loss(ping_response)
        
        ping_count = 1
        for latency in response_times:
            ping_time = time_when_command_started + timedelta(seconds=ping_count)
            if latency == 0:
                packet_dropped = True 
            else:
                packet_dropped = None
            my_data.loc[len(my_data.index)] = [ping_time, latency, packet_dropped]
            ping_count += 1
    
        if len(my_data) % save_frequency == 0: # Should be a multiple of four since ping responses have four values
            save_data(my_data, save_filename)

if __name__ == "__main__":
    start_time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    save_filename = f"ping_data_{start_time}.xlsx"
    

    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-t','--target', type=str, default='8.8.8.8', help='The host or IP address you want to test against.')
    parser.add_argument('-f', '--save-frequency', type=int, default=32, help='Number of pings to run between file saves. Should be a multiple of four.')
    parser.add_argument('-o', '--output-file', type=str, default=f"ping_data_{start_time}.xlsx", help='Output excel file.')
    args = parser.parse_args()
    if args.save_freqency % 4 != 0:
        raise ValueError(f"--save-frequency must be a multiple of four. You provided {args.save_frequency}.")

    main_loop(args.output_file, target=args.target, save_frequency=args.save_frequency)
    